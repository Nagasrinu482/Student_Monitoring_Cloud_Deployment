import datetime

from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, redirect

from student.forms import FacultyForm, StudentForm, AssignmentForm
from student.models import FacultyModel, StudentModel, AttendanceModel, MarksModel, AssignmentModel

import openpyxl
import os

def facultyregistration(request):

    if request.method == "POST":

        registrationForm = FacultyForm(request.POST)

        if registrationForm.is_valid():

            regModel = FacultyModel()
            regModel.name = registrationForm.cleaned_data["name"]
            regModel.email = registrationForm.cleaned_data["email"]
            regModel.mobile = registrationForm.cleaned_data["mobile"]
            regModel.department = registrationForm.cleaned_data["department"]
            regModel.username = registrationForm.cleaned_data["username"]
            regModel.password = registrationForm.cleaned_data["password"]

            user = FacultyModel.objects.filter(username=regModel.username).first()

            if user is not None:
                return render(request, 'facultyregistration.html', {"message": "User All Ready Exist"})
            else:
                try:
                    regModel.save()
                    return render(request, 'facultyregistration.html', {"message": "Faculty Added Successfully"})
                except:
                    return render(request, 'facultyregistration.html', {"message": "Registration Failed"})
        else:
            return render(request, 'facultyregistration.html', {"message": "Invalid Form"})

    return render(request, 'facultyregistration.html', {"message": "Invalid Request"})

def studentregistration(request):

    if request.method == "POST":

        registrationForm = StudentForm(request.POST)

        if registrationForm.is_valid():

            regModel = StudentModel()
            regModel.name = registrationForm.cleaned_data["name"]
            regModel.email = registrationForm.cleaned_data["email"]
            regModel.mobile = registrationForm.cleaned_data["mobile"]
            regModel.department = registrationForm.cleaned_data["department"]
            regModel.username = registrationForm.cleaned_data["username"]
            regModel.password = registrationForm.cleaned_data["password"]
            regModel.year = registrationForm.cleaned_data["year"]
            regModel.status ="no"

            user = StudentModel.objects.filter(username=regModel.username).first()

            if user is not None:
                return render(request, 'studentregistration.html', {"message": "User All Ready Exist"})
            else:
                try:
                    regModel.save()
                    return render(request, 'studentregistration.html', {"message": "Student Added Successfully"})
                except:
                    return render(request, 'studentregistration.html', {"message": "Registration Failed"})
        else:
            return render(request, 'studentregistration.html', {"message": "Invalid Form"})

    return render(request, 'studentregistration.html', {"message": "Invalid Request"})

#===============================================================================================
def deletestudent(request):
    student=request.GET['studentid']
    StudentModel.objects.get(id=student).delete()
    return render(request, 'students.html', {'students': StudentModel.objects.all()})

def deletefaculty(request):
    faculty=request.GET['facultyid']
    FacultyModel.objects.get(id=faculty).delete()
    return render(request, 'facultys.html', {'facultys': FacultyModel.objects.all()})

#===============================================================================================
def getfacultys(request):
    return render(request, "facultys.html", {"facultys":FacultyModel.objects.all()})

def getstudents(request):
    return render(request, "students.html", {"students":StudentModel.objects.all()})

#===============================================================================================
def login(request):

    uname = request.GET["username"]
    upass = request.GET["password"]
    type = request.GET["type"]

    if type in "admin":
        if uname == "admin" and upass == "admin":
            request.session['username'] = "admin"
            request.session['role'] = "admin"
            return render(request, "facultys.html", {"facultys":FacultyModel.objects.all()})
        else:
            return render(request, 'index.html', {"message": "Invalid Credentials"})

    if type in "student":
        student = StudentModel.objects.filter(username=uname, password=upass, status="yes").first()

        if student is not None:
            request.session['username'] = student.username
            request.session['department'] = student.department
            request.session['role'] = "student"
            return redirect(viewassignments)
        else:
            return render(request, 'index.html', {"message": "Invalid Username and Password"})

    if type in "faculty":
        faculty = FacultyModel.objects.filter(username=uname, password=upass).first()
        if faculty is not None:
            request.session['username']=faculty.username
            request.session['department'] = faculty.department
            request.session['role'] = "faculty"
            return render(request, 'addattendance.html')
        else:
            return render(request, 'index.html', {"message": "Invalid Username and Password"})

def activateAccount(request):

    username = request.GET['username']
    status=request.GET['status']

    StudentModel.objects.filter(username=username).update(status=status)
    return render(request, "students.html", {"students": StudentModel.objects.all()})

def logout(request):
    try:
        del request.session['username']
    except:
        pass
    return render(request, 'index.html', {})


def addattendance(request):

    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(file.name,file)
        uploaded_file_url = fs.url(filename)

    # Define variable to load the dataframe
    PROJECT_PATH = os.path.abspath(os.path.dirname(__name__))
    dataframe = openpyxl.load_workbook(PROJECT_PATH+uploaded_file_url)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # Iterate the loop to read the cell values
    attendedstudents=[]
    for row in range(1, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            attendedstudents.append(str(col[row].value))

    for student in StudentModel.objects.filter(department=request.session['department']):

        attendance = AttendanceModel()
        attendance.username = student.username
        attendance.date = datetime.datetime.now()

        if student.username in attendedstudents:
            attendance.isattended = "yes"
        else:
            attendance.isattended = "no"

        attendance.save()

    return redirect(viewattendanceaction)

def viewattendanceaction(request):

    percentagedict=dict()

    for student in StudentModel.objects.filter(department=request.session['department']):

        attendancedict = dict()

        count = AttendanceModel.objects.filter(username=student.username).count()

        if count > 0:

            for attendance in AttendanceModel.objects.filter(username=student.username):

                if attendance.isattended=="yes":
                    if attendance.username in attendancedict:
                        attendancedict[attendance.username] = attendancedict[attendance.username] + 1
                    else:
                        attendancedict[attendance.username] = 1
                else:
                    if attendance.username in attendancedict:
                        attendancedict[attendance.username] = attendancedict[attendance.username]
                    else:
                        attendancedict[attendance.username] = 0

            for key, val in attendancedict.items():
                percentagedict.update({key: (val / count) * 100})

    return render(request, "viewattendance.html", {"percentage":percentagedict})

def viewstudentattendance(request):
    return render(request, "viewstudentattendance.html",
                  {"attendances": AttendanceModel.objects.filter(username=request.GET['username'])})

def uploadmarks(request):

    if request.method == 'POST' and request.FILES['file']:
        file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(file.name,file)
        uploaded_file_url = fs.url(filename)

    # Define variable to load the dataframe
    PROJECT_PATH = os.path.abspath(os.path.dirname(__name__))
    dataframe = openpyxl.load_workbook(PROJECT_PATH+uploaded_file_url)

    # Define variable to read sheet
    dataframe1 = dataframe.active

    # Iterate the loop to read the cell values
    for row in range(1, dataframe1.max_row):
        marksrow=[]
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            marksrow.append(str(col[row].value))

        marksModel = MarksModel()
        marksModel.username = marksrow[0]
        marksModel.subjectcode = marksrow[1]
        marksModel.marks = marksrow[2]
        marksModel.save()

    return redirect(viewattendanceaction)

def viewmarks(request):
    return render(request, "viewmarks.html", {"students":StudentModel.objects.filter(department=request.session['department'])})
def viewmarksaction(request):
    return render(request, "viewstudentmarks.html", {"marks":MarksModel.objects.filter(username=request.GET['username'])})


def uploadassignment(request):
    return render(request, "uploadassignment.html",{})

def uploadassignmentaction(request):

    assignmentForm = AssignmentForm(request.POST, request.FILES)

    if assignmentForm.is_valid():
        department=request.session['department']
        document= assignmentForm.cleaned_data['document']
        subject = assignmentForm.cleaned_data['subject']
        AssignmentModel(department=department,subjectcode=subject,document=document).save()
        return redirect(viewassignments)
    return render(request, "uploadassignment.html", {"message":"invalid request"})

def viewassignments(request):

    assignments = []

    for assignment in AssignmentModel.objects.filter(department=request.session['department']):
        assignment.document = str(assignment.document).split('/')[1]
        assignments.append(assignment)

    return render(request, "viewassignments.html", {"assignments":assignments})